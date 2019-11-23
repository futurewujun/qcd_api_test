# -*- coding: utf-8 -*-
# 1：完成注册手机号码的初始化操作：修改Excel
# 第一种操作：利用Excel设置初始化手机号码，每次进行+1操作，以及变量替换。记得做数据演示。
# 第二种操作：每次从数据库里面查询最大的手机号码，在这个基础上加1（后期自己操作）
# 第三种操作：每次清除完这个手机号码相关的数据，进行垃圾数据重置操作。 当前时间戳生成手机号码 参数替换完成 ${变量名}

from openpyxl import load_workbook
from common import project_path
from common.read_config import ReadConfig
from common.do_re import Do_re
class DoExcel:
    '''完成测试数据的读取以及测试结果的写回'''

    def __init__(self,file_name,sheet_name):
        self.file_name=file_name    #Excel工作簿文件名或地址
        self.sheet_name=sheet_name

    def read_data(self,section):
        '''从Excel读取数据，有返回值'''
        #拿到配置文件的case_id
        case_id=ReadConfig(project_path.conf_path).get_data(section,'case_id')
        wb=load_workbook(self.file_name)
        st=wb[self.sheet_name]
        # tel=wb['tel'].cell(1,2).value
        tel=self.get_tel()
        # print(type(tel))
        #开始读取数据，
        test_data=[]
        for i in range(2,st.max_row+1): # 每行
            row_data={} # 每行的存到字典中，以每列的title作为key
            row_data['CaseId']=st.cell(i,1).value
            row_data['Module']=st.cell(i,2).value #第一列
            row_data['Title']=st.cell(i,3).value
            row_data['Url']=st.cell(i,4).value
            row_data['Method']=st.cell(i,5).value
            if st.cell(i,6).value.find('tel') != -1: #使用find函数，不存在返回-1，存在即!= -1
            # if 'tel' in st.cell(i, 6).value:    #也可用成员运算符
                row_data['Params']=st.cell(i,6).value.replace('tel',str(tel))
                self.update_tel(tel+1)
            else:
                row_data['Params'] = st.cell(i, 6).value
            # # 直接调用正则 查找替换
            # p = '#(.*?)#'
            # Do_re().do_re(p,)

            # print(st.cell(1,7).value)
            if st.cell(1,7).value == 'Sql': #判断是否有sql列--- 第一列是否是Sql
                row_data['Sql']=st.cell(i,7).value  #加的sql列
                row_data['ExpectedResult'] = st.cell(i, 8).value
            else:
                row_data['ExpectedResult']=st.cell(i,7).value

            test_data.append(row_data)
        wb.close()
        final_test_data=[]  #存最终的用例数据
        # 根据配置文件执行指定的用例
        if case_id == 1:    #等于1，执行所有用例
            final_test_data=test_data
        # 否则，如果是列表，获取列表里的数字执行指定的用例
        else:
            for i in case_id:    #遍历配置文件中case_id的值
                final_test_data.append(test_data[i-1]) #case_id=1,为test_data的第一条用例
        return final_test_data

    def update_tel(self,new_tel):
        '''更新tel的值'''
        wb=load_workbook(self.file_name)
        st=wb['tel']
        st.cell(1,2).value=new_tel
        wb.save(self.file_name)
        wb.close()
    def get_tel(self):
        '''获取tel的值'''
        wb=load_workbook(self.file_name)
        st=wb['tel']
        wb.close()
        tel=st.cell(1,2).value
        return tel  #返回tel的值

    def write_back(self,row,col,value):
        '''写回测试结果到Excel中'''
        wb=load_workbook(self.file_name)
        st=wb[self.sheet_name]
        st.cell(row,col).value=value
        # sheet.cell(3, 2, '作者：李白')
        wb.save(self.file_name)
        wb.close()

if __name__ == '__main__':
    file_name=project_path.case_path
    sheet_name='Login'
    test_data=DoExcel(file_name,sheet_name).read_data('LoginCase')
    # print(test_data)
    # 直接调用正则 查找替换
    p = '#(.*?)#'
    print(Do_re().do_re(p,str(test_data)))

